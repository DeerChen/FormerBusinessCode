#!/usr/bin/env python
# coding: utf-8

'''
@Author: Senkita
'''

import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side

pd.set_option('mode.chained_assignment', None)

class BackgroundProcessing:
    @classmethod
    def __init__(cls, date, order_start_num, waybill_num, flight_date, express_delivery, file_path):
        cls.date = date
        cls.order_start_num = order_start_num
        cls.waybill_num = waybill_num
        cls.flight_date = flight_date
        cls.express_delivery = express_delivery
        cls.file_path = file_path
        cls.raw_product_detail = cls.get_raw_product_detail()
        cls.product_detail = cls.organize_product_detail()
        cls.product_summary = cls.merge_product_detail()
        cls.costoms_template = cls.costoms_template()
        cls.yuantong_template = cls.yuantong_template()
        cls.yunda_template = cls.yunda_template()
    
    @staticmethod
    def set_style(worksheet, row_num, column_num):
        worksheet[chr(column_num+65)+str(row_num+1)].alignment = Alignment(horizontal='center', vertical='center')
        worksheet[chr(column_num+65)+str(row_num+1)].border = Border(
                    left=Side(style='thin',color='FF000000'),
                    right=Side(style='thin',color='FF000000'),
                    top=Side(style='thin',color='FF000000'),
                    bottom=Side(style='thin',color='FF000000')
                    )
        return worksheet

    @staticmethod
    def random_extract_sender():
        sender_list = [
            {'name': 'Abigail', 'phone': 6460594673, 'address': '2500 Nostrand Ave，Brooklyn, NY '},
            {'name': 'Albert', 'phone': 6460843858, 'address': '509 Myrtle Ave，Brooklyn, NY '},
            {'name': 'Allen', 'phone': 6460938616, 'address': '4508 5th Ave，Brooklyn, NY '},
            {'name': 'Amanda', 'phone': 3578213947, 'address': '64 Frost St，Brooklyn, NY'},
            {'name': 'Andrew', 'phone': 2121971196, 'address': '1327 E 99th St，Brooklyn, NY'},
            {'name': 'Ann ', 'phone': 3578219219, 'address': '7325 Woodhaven Blvd，Ridgewood, NY'},
            {'name': 'Barbara', 'phone': 3578217356, 'address': '2029 76th St，Brooklyn, NY'},
            {'name': 'Bruce', 'phone': 6460693357, 'address': '231 Norman Avenue #104，Brooklyn, NY '},
            {'name': 'Carl', 'phone': 6460831978, 'address': '298 Bedford Ave，Brooklyn, NY'},
            {'name': 'Charles', 'phone': 3578214287, 'address': '52 Quentin Rd，Brooklyn, NY'},
            {'name': 'Charlie', 'phone': 6460914073, 'address': '1825 Church Ave，Brooklyn, NY '},
            {'name': 'Dan', 'phone': 3578213939, 'address': '751 Knickerbocker Ave，Brooklyn, NY'},
            {'name': 'David', 'phone': 2121147084, 'address': '1790 Stillwell Ave，Brooklyn, NY'},
            {'name': 'Dean', 'phone': 3578215296, 'address': '960 Cypress Ave，Ridgewood, NY '},
            {'name': 'Doris', 'phone': 2121429314, 'address': '57 Waterbury St，Brooklyn, NY'},
            {'name': 'Ethan', 'phone': 3578218648, 'address': '1836 Rockaway Pkwy，Brooklyn, NY'},
            {'name': 'Fred', 'phone': 2121968170, 'address': '64-66 Myrtle Ave，Glendale, NY'},
            {'name': 'George', 'phone': 3578214729, 'address': '301 Vermont St，Brooklyn, NY'},
            {'name': 'Grace', 'phone': 6460528046, 'address': '2147 70th St，Brooklyn, NY'},
            {'name': 'Hailey', 'phone': 2121532745, 'address': '931 Thomas S Boyland St # A，Brooklyn, NY'},
            {'name': 'Hannah', 'phone': 2121160656, 'address': '1764 E 18th St，Brooklyn, NY '},
            {'name': 'Helen', 'phone': 6460782073, 'address': '28 Crosby Ave，Brooklyn, NY'},
            {'name': 'Howard', 'phone': 2121035830, 'address': '5001 6th Ave，Brooklyn, NY'},
            {'name': 'Isabella', 'phone': 6460521738, 'address': '2236 Nostrand Ave，Brooklyn, NY'},
            {'name': 'Jacob', 'phone': 2121710787, 'address': '4207 Avenue J，Brooklyn, NY'},
            {'name': 'James', 'phone': 3578215762, 'address': '8686 Bay Pkwy，Brooklyn, NY'},
            {'name': 'Jennifer', 'phone': 2121873242, 'address': '1744 Ocean Pkwy，Brooklyn, NY '},
            {'name': 'Jessica', 'phone': 3578212902, 'address': '1112 Bedford Ave，Brooklyn, NY '},
            {'name': 'Johnson', 'phone': 3578217887, 'address': '645 Manhattan Ave，Brooklyn, NY'},
            {'name': 'Joy', 'phone': 3578212362, 'address': '12-09 Jackson Ave，Long Island City, NY'},
            {'name': 'Kevin', 'phone': 3578219702, 'address': '330 Powell St，Brooklyn, NY'},
            {'name': 'Lillian', 'phone': 6460453605, 'address': '278 Greenpoint Ave，Brooklyn, NY'},
            {'name': 'Lily', 'phone': 6460080979, 'address': '6819 Fresh Pond Rd，Ridgewood, NY '},
            {'name': 'Lisa', 'phone': 6460804697, 'address': '195 Jamaica Ave，Brooklyn, NY'},
            {'name': 'Madison', 'phone': 3578211542, 'address': '1101 Avenue S，Brooklyn, NY '},
            {'name': 'Malcolm', 'phone': 2121205605, 'address': '128 Pierrepont St，Brooklyn, NY'},
            {'name': 'Maria', 'phone': 6460453605, 'address': '2050 86th St Brooklyn, NY '},
            {'name': 'Mark', 'phone': 6460943516, 'address': '209 Avenue P，Brooklyn, NY'},
            {'name': 'Melody', 'phone': 2121689171, 'address': '3038 Atlantic Ave，Brooklyn, NY '},
            {'name': 'Niki', 'phone': 3578212080, 'address': '61 15 Metropolitan Ave，Ridgewood, NY.'},
            {'name': 'Olivia', 'phone': 3578214424, 'address': '2474 Ocean Ave，Brooklyn, NY'},
            {'name': 'Paul', 'phone': 3578211427, 'address': '42 Avenue O，Brooklyn, NY'},
            {'name': 'Peter', 'phone': 3578215762, 'address': '114 Greenpoint Ave，Brooklyn, NY '},
            {'name': 'Rebecca', 'phone': 3578219458, 'address': '683 Washington Ave，Brooklyn, NY'},
            {'name': 'Richard', 'phone': 3578216138, 'address': '1801 Stillwell Ave，Brooklyn, NY '},
            {'name': 'Rita', 'phone': 3578218937, 'address': '1014 Fulton St，Brooklyn, NY'},
            {'name': 'Robert', 'phone': 3578217374, 'address': '136 Metropolitan Ave，Brooklyn, NY '},
            {'name': 'Sam', 'phone': 2121510161, 'address': '2818 Coney Island Ave，Brooklyn, NY'},
            {'name': 'Samantha', 'phone': 3578214288, 'address': '4018 Glenwood Rd，Brooklyn, NY'},
            {'name': 'Scott', 'phone': 3578219702, 'address': '60 S 2nd St，Brooklyn, NY'},
            {'name': 'Shelly', 'phone': 3578212852, 'address': '910 Manhattan Ave，Brooklyn, NY'},
            {'name': 'Shirley', 'phone': 2121358497, 'address': '2 Havemeyer St，Brooklyn, NY '},
            {'name': 'Sophia', 'phone': 6460949435, 'address': '135 India St，Brooklyn, NY'},
            {'name': 'Sunny', 'phone': 2121253722, 'address': '6722 Fort Hamilton Pkwy，Brooklyn, NY'},
            {'name': 'Teresa', 'phone': 6460762582, 'address': '583 Vanderbilt Ave，Brooklyn, NY'},
            {'name': 'Vincent', 'phone': 2121054370, 'address': '748 Myrtle Ave，Brooklyn, NY'},
            {'name': 'Vivian', 'phone': 3578211213, 'address': '256 McGuinness Blvd，Brooklyn, NY'},
            {'name': 'Wendy', 'phone': 3578214424, 'address': '1424 Avenue J，Brooklyn, NY '}
            ]
        sender = random.choice(sender_list)
        return sender['name'], sender['phone'], sender['address']

    @classmethod
    def get_raw_product_detail(cls):
        product_detail_file = pd.read_csv(cls.file_path, encoding='gbk', low_memory=False)
        raw_product_detail = pd.concat([product_detail_file, pd.DataFrame(columns=['克数', '规格数量'])], axis=1)
        for i in range(raw_product_detail.shape[0]):
            specification = str(raw_product_detail['规格名称'][i])
            raw_product_detail['克数'][i] = specification.split('*')[0]
            try:
                raw_product_detail['规格数量'][i] = specification.split('*')[1]
            except:
                raw_product_detail['规格数量'][i] = 1
        return raw_product_detail

    @classmethod
    def organize_product_detail(cls):
        product_detail = pd.DataFrame(columns=[
            '单据编号',
            '系统发货单号',
            '序号',
            '落地配单号/分运单号',
            '配货单号',
            '后4位',
            '发件人(英文)/寄件人',
            '发件人电话/寄件人电话',
            '发件人地址/寄件人地址',
            '英文发件人城市/寄件人城市',
            '收货人/收件人',
            '收货人电话/收件人电话',
            '身份证号码/身份证',
            '收货人城市/省份',
            '城市',
            '区县',
            '收货人地址/详细地址(包含省市区县)',
            '税号',
            '品牌',
            '原始品名',
            '商品名称/货物品名',
            '商品规格、型号/规格型号',
            '数量',
            '实重/净重(kg)',
            '重量',
            '毛重/毛重(kg)',
            '申报单价',
            '生产厂商',
            '快件运营单位',
            '英文经停城市',
            '发件人国别',
            '收货人国别',
            '商品名称(英文)',
            '件数',
            '单位',
            '包装种类',
            '行邮税号',
            '申报总价',
            '币别',
            '国别',
            '段数',
            '克数',
            '材质',
            '寄件人所在省',
            '寄件人所在区',
            '卖家备注',
            '日期备注'
            ])
        for i in range(cls.raw_product_detail.shape[0]):
            num = cls.raw_product_detail['数量'][i] * int(cls.raw_product_detail['规格数量'][i])
            try:
                total_price = cls.raw_product_detail['数量'][i] * int(cls.raw_product_detail['海关定价'][i])
            except:
                total_price = 0

            product_detail.loc[i, '单据编号'] = cls.raw_product_detail['单据编号'][i]
            product_detail.loc[i, '系统发货单号'] = cls.raw_product_detail['单据编号'][i]
            product_detail.loc[i, '发件人(英文)/寄件人'], product_detail.loc[i, '发件人电话/寄件人电话'], product_detail.loc[i, '发件人地址/寄件人地址'] = cls.random_extract_sender()
            product_detail.loc[i, '收货人/收件人'] = cls.raw_product_detail['收货人'][i]
            product_detail.loc[i, '收货人电话/收件人电话'] = cls.raw_product_detail['收货人手机'][i]
            product_detail.loc[i, '身份证号码/身份证'] = cls.raw_product_detail['身份证号'][i]            
            product_detail.loc[i, '收货人地址/详细地址(包含省市区县)'] = cls.raw_product_detail['收货地址'][i]
            product_detail.loc[i, '商品规格、型号/规格型号'] = '{}|{}{}|{}{}/{}|{}{}'.format(
                cls.raw_product_detail['类目'][i],
                cls.raw_product_detail['英文牌子名称'][i],
                cls.raw_product_detail['品牌'][i],
                cls.raw_product_detail['配方'][i],
                cls.raw_product_detail['克数'][i],
                cls.raw_product_detail['商品单位'][i],
                cls.raw_product_detail['年龄'][i],
                cls.raw_product_detail['段数'][i]
                ).replace('nan', '')
            product_detail.loc[i, '数量'] = num
            product_detail.loc[i, '实重/净重(kg)'] = cls.raw_product_detail['总重量'][i]
            product_detail.loc[i, '重量'] = cls.raw_product_detail['总重量'][i]
            product_detail.loc[i, '毛重/毛重(kg)'] = cls.raw_product_detail['总重量'][i]
            product_detail.loc[i, '申报单价'] = cls.raw_product_detail['海关定价'][i]
            product_detail.loc[i, '商品名称(英文)'] = '{} {}*{:.0f}'.format(
                cls.raw_product_detail['英文牌子名称'][i],
                cls.raw_product_detail['英文类别名称'][i],
                num
                ).replace('nan', '')
            product_detail.loc[i, '单位'] = cls.raw_product_detail['商品单位'][i]
            product_detail.loc[i, '申报总价'] = total_price
            product_detail.loc[i, '卖家备注'] = cls.raw_product_detail['卖家备注'][i]
            product_detail.loc[i, '日期备注'] = cls.raw_product_detail['收货人电话'][i]

            try:
                product_detail.loc[i, '收货人城市/省份'] = cls.raw_product_detail['收货地址'][i].split(' ')[0]
                product_detail.loc[i, '城市'] = cls.raw_product_detail['收货地址'][i].split(' ')[1]
                product_detail.loc[i, '区县'] = cls.raw_product_detail['收货地址'][i].split(' ')[2]
            except:
                pass
            
            if cls.raw_product_detail['类目'][i] in ('奶粉', '液态奶'):
                if cls.raw_product_detail['段数'][i] == '成人':
                    product_detail.loc[i, '商品名称/货物品名'] = '{}{}{}{}*{:.0f}{}'.format(
                        cls.raw_product_detail['英文牌子名称'][i],
                        cls.raw_product_detail['品牌'][i],
                        cls.raw_product_detail['段数'][i],
                        cls.raw_product_detail['类目'][i],
                        num,
                        cls.raw_product_detail['商品单位'][i]
                        ).replace('nan', '')
                else:
                    product_detail.loc[i, '商品名称/货物品名'] = '{}{}{}{}{}*{:.0f}{}'.format(
                        cls.raw_product_detail['英文牌子名称'][i],
                        cls.raw_product_detail['品牌'][i],
                        cls.raw_product_detail['年龄'][i],
                        cls.raw_product_detail['类目'][i],
                        cls.raw_product_detail['段数'][i],
                        num,
                        cls.raw_product_detail['商品单位'][i]
                        ).replace('nan', '')
            else:
                product_detail.loc[i, '商品名称/货物品名'] = '{}*{:.0f}{}'.format(
                    cls.raw_product_detail['类目'][i],
                    num,
                    cls.raw_product_detail['商品单位'][i]
                    ).replace('nan', '')
            
            if cls.raw_product_detail['配方'][i] == '小安素':
                product_detail.loc[i, '商品名称/货物品名'] = '{}{}{}{}*{:.0f}{}'.format(
                    cls.raw_product_detail['英文牌子名称'][i],
                    cls.raw_product_detail['品牌'][i],
                    cls.raw_product_detail['配方'][i],
                    cls.raw_product_detail['段数'][i],
                    num,
                    cls.raw_product_detail['商品单位'][i]
                    ).replace('nan', '')

            if pd.isna(cls.raw_product_detail['克数'][i]):
                product_detail.loc[i, '克数'] = ''
            else:
                product_detail.loc[i, '克数'] = cls.raw_product_detail['克数'][i]

            if pd.isna(cls.raw_product_detail['段数'][i]):
                product_detail.loc[i, '段数'] = ''
            else:
                product_detail.loc[i, '段数'] = cls.raw_product_detail['段数'][i]
            
            if pd.isna(cls.raw_product_detail['英文牌子名称'][i]):
                product_detail.loc[i, '生产厂商'] = ''
                product_detail.loc[i, '品牌'] = ''
            else:
                product_detail.loc[i, '生产厂商'] = cls.raw_product_detail['英文牌子名称'][i]
                product_detail.loc[i, '品牌'] = cls.raw_product_detail['英文牌子名称'][i]
            
            if pd.isna(cls.raw_product_detail['收货人电话'][i]):
                product_detail.loc[i, '原始品名'] = '{}*{:.0f}'.format(
                    cls.raw_product_detail['商品名称'][i],
                    cls.raw_product_detail['数量'][i]
                    )
            else:
                product_detail.loc[i, '原始品名'] = '{}*{:.0f}({})'.format(
                    cls.raw_product_detail['商品名称'][i],
                    cls.raw_product_detail['数量'][i],
                    cls.raw_product_detail['收货人电话'][i]
                    )

        product_detail['英文发件人城市/寄件人城市'] = 'NewYork'
        product_detail['快件运营单位'] = '厦门东港国际运输有限公司'
        product_detail['英文经停城市'] = 'NewYork'
        product_detail['寄件人所在省'] = 'NewYork'
        product_detail['寄件人所在区'] = 'NewYork'
        product_detail['发件人国别'] = '美国'
        product_detail['收货人国别'] = '中国'
        product_detail['件数'] = 1
        product_detail['包装种类'] = '纸箱'
        product_detail['币别'] = 'RMB'
        product_detail['国别'] = '美国'
        return product_detail

    @classmethod
    def merge_product_detail(cls):
        product_detail_cut_version = cls.product_detail.drop_duplicates(subset='单据编号').drop([
            '原始品名',
            '商品名称/货物品名',
            '商品名称(英文)',
            '商品规格、型号/规格型号',
            '数量',
            '实重/净重(kg)',
            '重量',
            '毛重/毛重(kg)',
            '申报单价',
            '申报总价',
            '生产厂商',
            '段数',
            '克数',
            '品牌'
            ], axis=1)
        product_detail_cut_version.set_index('单据编号', inplace=True)

        product_original_name = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['原始品名'])), columns=['原始品名'])
        product_name = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['商品名称/货物品名'])), columns=['商品名称/货物品名'])
        product_en_name = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['商品名称(英文)'])), columns=['商品名称(英文)'])
        product_specification = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['商品规格、型号/规格型号'])), columns=['商品规格、型号/规格型号'])
        product_num = cls.product_detail.groupby(by='单据编号')['数量'].sum()
        product_net_weight = cls.product_detail.groupby(by='单据编号')['实重/净重(kg)'].sum()
        product_weight = cls.product_detail.groupby(by='单据编号')['重量'].sum()
        product_gross_weight = cls.product_detail.groupby(by='单据编号')['毛重/毛重(kg)'].sum()
        product_unit_price = cls.product_detail.groupby(by='单据编号')['申报单价'].sum()
        product_total_price = cls.product_detail.groupby(by='单据编号')['申报总价'].sum()
        product_manufacturer = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['生产厂商'])), columns=['生产厂商'])
        product_step = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['段数'])), columns=['段数'])
        product_grams = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['克数'])), columns=['克数'])
        product_brand = pd.DataFrame(cls.product_detail.groupby(by='单据编号').apply(lambda x: ','.join(x['品牌'])), columns=['品牌'])
        product_summary = pd.concat([
            product_detail_cut_version,
            product_original_name,
            product_name,
            product_en_name,
            product_specification,
            product_num,
            product_net_weight,
            product_weight,
            product_gross_weight,
            product_unit_price,
            product_total_price,
            product_manufacturer,
            product_step,
            product_grams,
            product_brand
            ], axis=1, sort=False)
        return product_summary

    @classmethod
    def costoms_template(cls):
        costoms_template = pd.concat([
            cls.product_summary['序号'],
            cls.product_summary['落地配单号/分运单号'],
            cls.product_summary['配货单号'],
            cls.product_summary['后4位'],
            cls.product_summary['系统发货单号'],
            cls.product_summary['快件运营单位'],
            cls.product_summary['发件人(英文)/寄件人'],
            cls.product_summary['发件人地址/寄件人地址'],
            cls.product_summary['英文发件人城市/寄件人城市'],
            cls.product_summary['英文经停城市'],
            cls.product_summary['发件人国别'],
            cls.product_summary['发件人电话/寄件人电话'],
            cls.product_summary['收货人/收件人'],
            cls.product_summary['收货人国别'],
            cls.product_summary['收货人电话/收件人电话'],
            cls.product_summary['身份证号码/身份证'],
            cls.product_summary['收货人地址/详细地址(包含省市区县)'],
            cls.product_summary['收货人城市/省份'],
            cls.product_summary['城市'],
            cls.product_summary['区县'],
            cls.product_summary['实重/净重(kg)'],
            cls.product_summary['重量'],
            cls.product_summary['原始品名'],
            cls.product_summary['商品名称/货物品名'],
            cls.product_summary['商品名称(英文)'],
            cls.product_summary['商品规格、型号/规格型号'],
            cls.product_summary['件数'],
            cls.product_summary['数量'],
            cls.product_summary['毛重/毛重(kg)'],
            cls.product_summary['单位'],
            cls.product_summary['包装种类'],
            cls.product_summary['行邮税号'],
            cls.product_summary['申报单价'],
            cls.product_summary['申报总价'],
            cls.product_summary['币别'],
            cls.product_summary['生产厂商'],
            cls.product_summary['国别'],
            cls.product_summary['段数'],
            cls.product_summary['克数'],
            cls.product_summary['材质'],
            cls.product_summary['卖家备注'],
            cls.product_summary['日期备注']
            ], axis=1, sort=False)
        costoms_template.sort_values(by=['原始品名'], inplace=True)
        costoms_template.rename(columns={
            '落地配单号/分运单号': '落地配单号',
            '发件人(英文)/寄件人': '发件人',
            '发件人地址/寄件人地址': '发件人地址',
            '英文发件人城市/寄件人城市': '英文发件人城市',
            '发件人电话/寄件人电话': '发件人电话',
            '收货人/收件人': '收货人',
            '收货人电话/收件人电话': '收货人电话',
            '身份证号码/身份证': '身份证号码',
            '收货人地址/详细地址(包含省市区县)': '收货人地址',
            '收货人城市/省份': '收货人城市',
            '城市': '市',
            '区县': '区',
            '实重/净重(kg)': '实重',
            '商品名称/货物品名': '商品名称',
            '商品规格、型号/规格型号': '商品规格、型号',
            '毛重/毛重(kg)': '毛重'
            }, inplace=True)
        for i in range(costoms_template.shape[0]):
            costoms_template['序号'][i] = int(i + 1)
            costoms_template['配货单号'][i] = '2K{}{:0>4d}NY'.format(cls.date[2:], int(cls.order_start_num)+i)
        return costoms_template

    @classmethod
    def yuantong_template(cls):
        product_summary_minial_version = cls.product_summary.drop([
            '序号',
            '落地配单号/分运单号',
            '配货单号',
            '后4位',
            '系统发货单号',
            '快件运营单位',
            '发件人(英文)/寄件人',
            '发件人地址/寄件人地址',
            '英文发件人城市/寄件人城市',
            '英文经停城市',
            '发件人国别',
            '发件人电话/寄件人电话',
            '收货人/收件人',
            '收货人国别',
            '收货人电话/收件人电话',
            '身份证号码/身份证',
            '收货人地址/详细地址(包含省市区县)',
            '收货人城市/省份',
            '城市',
            '区县',
            '实重/净重(kg)',
            '重量',
            '原始品名',
            '商品名称/货物品名',
            '商品名称(英文)',
            '商品规格、型号/规格型号',
            '件数',
            '数量',
            '毛重/毛重(kg)',
            '单位',
            '包装种类',
            '行邮税号',
            '申报单价',
            '申报总价',
            '币别',
            '生产厂商',
            '国别',
            '段数',
            '克数',
            '材质',
            '卖家备注',
            '日期备注'
            ], axis=1)
        costoms_template_minial_version = cls.costoms_template.rename(columns={
            '落地配单号': '分运单号',
            '系统发货单号': '单据编号',
            '发件人': '寄件人',
            '发件人电话': '寄件人电话',
            '发件人地址': '寄件人地址',
            '英文发件人城市': '寄件人城市',
            '收货人': '收件人',
            '收货人电话': '收件人电话',
            '身份证号码': '身份证',
            '收货人城市': '省份',
            '市': '城市',
            '区': '区县',
            '收货人地址': '详细地址(包含省市区县)',
            '商品名称': '货物品名',
            '商品规格、型号': '规格型号',
            '实重': '净重(kg)',
            '毛重': '毛量(kg)'
            })
        costoms_template_minial_version.set_index('单据编号', inplace=True)
        complete_summary = pd.concat([
            product_summary_minial_version,
            costoms_template_minial_version
            ], axis=1, sort=False)
        yuantong_template = pd.concat([
            complete_summary['序号'],
            complete_summary['分运单号'],
            complete_summary['配货单号'],
            complete_summary['寄件人'],
            complete_summary['寄件人电话'],
            complete_summary['寄件人地址'],
            complete_summary['寄件人城市'],
            complete_summary['收件人'],
            complete_summary['收件人电话'],
            complete_summary['身份证'],
            complete_summary['省份'],
            complete_summary['城市'],
            complete_summary['区县'],
            complete_summary['详细地址(包含省市区县)'],
            complete_summary['税号'],
            complete_summary['品牌'],
            complete_summary['货物品名'],
            complete_summary['克数'],
            complete_summary['数量'],
            complete_summary['净重(kg)'],
            complete_summary['毛量(kg)'],
            complete_summary['申报总价'],
            complete_summary['生产厂商'],
            ], axis=1, sort=False)
        yuantong_template.rename(columns={
            '克数': '规格型号',
            '申报总价': '申报单价'
            }, inplace=True)
        yuantong_template.sort_values(by=['序号'], inplace=True)
        return yuantong_template

    @classmethod
    def generate_yuantong_template(cls):
        name = '{}.{}圆通单号申请.xlsx'.format(cls.date[4:6], cls.date[6:])
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = '总运单号'
        ws.cell(1, 2).value = cls.waybill_num
        ws.cell(2, 1).value = '航班'
        ws.cell(2, 2).value = 'MF850'
        ws.cell(3, 1).value = '航班日期'
        ws.cell(3, 2).value = datetime.strftime(cls.flight_date, '%Y%m%d')
        ws.cell(4, 1).value = '进境口岸'
        ws.cell(4, 2).value = '福州'
        ws.cell(5, 1).value = '进境日期'
        ws.cell(5, 2).value = datetime.strftime(cls.flight_date-timedelta(days=1), '%Y%m%d')
        ws.cell(6, 1).value = '启运国'
        ws.cell(6, 2).value = '美国'
        ws.cell(7, 1).value = '目的国'
        ws.cell(7, 2).value = '晋江'
        ws.merge_cells('A8:C8')
        ws.cell(8, 1).value = '运单信息'
        ws.merge_cells('D8:G8')
        ws.cell(8, 4).value = '寄件人信息'
        ws.merge_cells('H8:N8')
        ws.cell(8, 8).value = '收件人信息'
        ws.merge_cells('O8:W8')
        ws.cell(8, 15).value = '货物信息'
        ws.cell(8, 1)
        for r in dataframe_to_rows(cls.yuantong_template, index=False, header=True):
            ws.append(r)
        for i in range(ws.max_row):
            for j in range(ws.max_column):
                if i >= 7:
                    ws = cls.set_style(ws, i, j)
                elif ws.cell(i+1, j+1).value is not None:
                    ws = cls.set_style(ws, i, j)
        wb.save(name)

    @classmethod
    def yunda_template(cls):
        product_summary_minial_version = cls.product_summary.drop([
            '序号',
            '落地配单号/分运单号',
            '配货单号',
            '后4位',
            '系统发货单号',
            '快件运营单位',
            '发件人(英文)/寄件人',
            '发件人地址/寄件人地址',
            '英文发件人城市/寄件人城市',
            '英文经停城市',
            '发件人国别',
            '发件人电话/寄件人电话',
            '收货人/收件人',
            '收货人国别',
            '收货人电话/收件人电话',
            '身份证号码/身份证',
            '收货人地址/详细地址(包含省市区县)',
            '收货人城市/省份',
            '城市',
            '区县',
            '实重/净重(kg)',
            '重量',
            '原始品名',
            '商品名称/货物品名',
            '商品名称(英文)',
            '商品规格、型号/规格型号',
            '件数',
            '数量',
            '毛重/毛重(kg)',
            '单位',
            '包装种类',
            '行邮税号',
            '申报单价',
            '申报总价',
            '币别',
            '生产厂商',
            '国别',
            '段数',
            '克数',
            '材质',
            '卖家备注',
            '日期备注'
            ], axis=1)
        product_summary_minial_version.rename(columns={
            '品牌': '商品品牌'
            }, inplace=True)
        costoms_template_minial_version = cls.costoms_template.rename(columns={
            '配货单号': '原单号',
            '系统发货单号': '单据编号',
            '发件人国别': '寄件人国别',
            '发件人': '寄件人',
            '发件人电话': '寄件人电话',
            '发件人地址': '寄件人地址',
            '英文发件人城市': '寄件人所在市',
            '收货人国别': '收件人国别',
            '收货人': '收件人',
            '收货人电话': '收件人电话',
            '身份证号码': '身份证',
            '收货人城市': '收件人所在省',
            '市': '收件人所在市',
            '区': '收件人所在区',
            '收货人地址': '详细地址(包含省市区县)',
            '商品规格、型号': '商品规格型号',
            '实重': '净重(kg)',
            '毛重': '毛量(kg)',
            '申报单价': '单价(￥)',
            '国别': '原产国'
            })
        costoms_template_minial_version.set_index('单据编号', inplace=True)
        complete_summary = pd.concat([
            product_summary_minial_version,
            costoms_template_minial_version
            ], axis=1, sort=False)
        yunda_template = pd.concat([
            complete_summary['序号'],
            complete_summary['原单号'],
            complete_summary['寄件人国别'],
            complete_summary['寄件人'],
            complete_summary['寄件人电话'],
            complete_summary['寄件人所在省'],
            complete_summary['寄件人所在市'],
            complete_summary['寄件人所在区'],
            complete_summary['寄件人地址'],
            complete_summary['收件人国别'],
            complete_summary['收件人'],
            complete_summary['收件人电话'],
            complete_summary['身份证'],
            complete_summary['收件人所在省'],
            complete_summary['收件人所在市'],
            complete_summary['收件人所在区'],
            complete_summary['详细地址(包含省市区县)'],
            complete_summary['商品品牌'],
            complete_summary['商品名称'],
            complete_summary['商品规格型号'],
            complete_summary['数量'],
            complete_summary['毛量(kg)'],
            complete_summary['净重(kg)'],
            complete_summary['单价(￥)'],
            complete_summary['原产国'],
            ], axis=1, sort=False)
        yunda_template.sort_values(by=['序号'], inplace=True)
        return yunda_template

    @classmethod
    def generate_yunda_template(cls):
        name = '{}.{}韵达单号申请.xlsx'.format(cls.date[4:6], cls.date[6:])
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = '电商客户名称'
        ws.cell(1, 2).value = '厦门东港国际有限公司'
        ws.cell(2, 1).value = '总运单号/提单号'
        ws.cell(2, 2).value = cls.waybill_num
        ws.cell(3, 1).value = '起运国家'
        ws.cell(3, 2).value = '美国'
        ws.cell(4, 1).value = '航班航次号'
        ws.cell(4, 2).value = 'MF850'
        ws.cell(5, 1).value = '进境日期'
        ws.cell(5, 2).value = datetime.strftime(cls.flight_date-timedelta(days=1), '%Y-%m-%d')
        ws.merge_cells('A6:B6')
        ws.cell(6, 1).value = '运单信息'
        ws.merge_cells('C6:I6')
        ws.cell(6, 3).value = '寄件人信息'
        ws.merge_cells('J6:Q6')
        ws.cell(6, 10).value = '收件人信息'
        ws.merge_cells('R6:Y6')
        ws.cell(6, 18).value = '货物信息'
        ws.cell(6, 1)
        for r in dataframe_to_rows(cls.yunda_template, index=False, header=True):
            ws.append(r)
        for i in range(ws.max_row):
            for j in range(ws.max_column):
                if i >= 5:
                    ws = cls.set_style(ws, i, j)
                elif ws.cell(i+1, j+1).value is not None:
                    ws = cls.set_style(ws, i, j)
        wb.save(name)

    @classmethod
    def generate_template(cls):
        cls.costoms_template.to_excel('{}.{}晋江清单.xlsx'.format(cls.date[4:6], cls.date[6:]), index=False)
        if cls.express_delivery == '圆通速递':
            cls.generate_yuantong_template()
        elif cls.express_delivery == '韵达速递':
            cls.generate_yunda_template()
