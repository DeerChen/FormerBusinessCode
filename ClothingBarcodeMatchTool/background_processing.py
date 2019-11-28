#!/usr/bin/env python
# coding: utf-8

'''
@Author: Senkita
'''

import pandas as pd
import numpy as np
from openpyxl import Workbook
import re

pd.set_option('mode.chained_assignment', None)

class BackgroundProcessing:
    @classmethod
    def __init__(cls, file_path, match_file_path, barcode_choice, tag_choice):
        cls.file = pd.read_excel(file_path)
        cls.match_file = pd.read_excel(match_file_path)
        cls.barcode_choice = barcode_choice
        cls.tag_choice = tag_choice
          
    @classmethod
    def match_barcode(cls):
        cls.file['条码数量'] = np.nan
        cls.file['吊牌数量'] = np.nan
        not_found_barcode = []
        
        for barcode in cls.match_file['商品条码']:
            file_index = cls.file[cls.file['条码']==barcode].index
            match_index = cls.match_file[cls.match_file['商品条码']==barcode].index
            
            if file_index.size == 0:
                not_found_barcode.append((barcode, int(cls.match_file['数量'][match_index])))
            else:
                if cls.barcode_choice == True:
                    cls.file['条码数量'][file_index] = cls.match_file['数量'][match_index]
                if cls.tag_choice == True:
                    cls.file['吊牌数量'][file_index] = cls.match_file['数量'][match_index]

        cls.file.to_excel('男装条码匹配.xlsx', index=False)
        barcode_count = '共匹配{}个条码'.format(cls.file['条码数量'].sum())
        tag_count = '共匹配{}个吊牌'.format(cls.file['吊牌数量'].sum())

        if len(not_found_barcode) != 0:
            wb = Workbook()
            ws = wb.active
            ws.cell(1, 1).value = '未匹配条码'
            ws.cell(1, 2).value = '数量'
            
            for i in range(len(not_found_barcode)):
                ws.cell(i + 2, 1).value = not_found_barcode[i][0]
                ws.cell(i + 2, 2).value = not_found_barcode[i][1]

            wb.save('未匹配条码.xlsx')

        return barcode_count, tag_count
